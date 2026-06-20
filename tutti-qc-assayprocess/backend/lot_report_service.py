"""Generate Tutti lot output workbooks from RDS AssayProcess records."""

from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell
import psycopg2.extras

from all_batch_service import (
    CONTROL_LEVELS,
    _cv,
    _find_spec,
    _mean,
    _number,
    _parse_level_thresholds,
    _parse_range,
    _parse_threshold,
    _round,
    parse_production_date,
)
from app_config import SPEC_DB_PATH
from query_service import _get_conn
from baseline_service import _fetch_batch_info


ROOT = Path(__file__).resolve().parents[1]
EXCEL_DATA_DIR = ROOT / "Excel-data"
TEMPLATE_PATH = EXCEL_DATA_DIR / "templet" / "Tutti 預設表格-h.xlsx"
CUTOFF_DATE = date(2026, 6, 10)

CONTROL_PATIENT_IDS = ("control-1", "control-2", "control-3", "control-4")
SPECIES_SHEETS = {
    "canine": "Canine",
    "feline": "Feline",
    "equine": "Equine",
}

RESULT_COLUMNS = {
    "Control-1": {"tea": 6, "assigned": 7, "upper": 8, "lower": 9, "values": list(range(10, 19)), "mean": 60, "bias": 71, "cv": 82},
    "Control-2": {"tea": 84, "assigned": 85, "upper": 86, "lower": 87, "values": list(range(88, 109)), "mean": 138, "bias": 149, "cv": 160},
    "Control-3": {"tea": 162, "assigned": 163, "upper": 164, "lower": 165, "values": list(range(166, 177)), "mean": 216, "bias": 227, "cv": 238},
    "Control-4": {"tea": 240, "assigned": 241, "upper": 242, "lower": 243, "values": list(range(244, 254)), "mean": 294, "bias": 305, "cv": 316},
}

CONTROL_SUMMARY_ROWS = {
    "Control-1": (9, 10, 11),
    "Control-2": (12, 13, 14),
    "Control-3": (15, 16, 17),
    "Control-4": (18, 19, 20),
}

SUMMARY_MARKER_COLUMNS = [5, 7, 9, 11, 13, 15, 17, 19, 21, 23]
SUMMARY_OD_COLUMNS = [27, 28, 29, 30, 31, 32, 33, 34, 35, 36]
DETAIL_START_ROW = 36
DETAIL_MAX_ROWS = 55


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _normalized_lot_code(value: Any) -> str:
    return _clean_text(value).replace("_", "")


def _lot_suffix(value: str) -> int | None:
    match = re.search(r"(\d{2})$", value or "")
    return int(match.group(1)) if match else None


def _lot_group_key(lot_code: str) -> str:
    normalized = _normalized_lot_code(lot_code)
    if len(normalized) == 12 and normalized.isdigit():
        return normalized[1:]
    return normalized


def _production_date_from_lot_code(lot_code: str, fallback: str = "") -> date | None:
    normalized = _normalized_lot_code(lot_code)
    if len(normalized) == 12 and normalized.isdigit():
        try:
            return datetime.strptime(normalized[4:10], "%y%m%d").date()
        except ValueError:
            pass
    return parse_production_date("", fallback)


def _include_lot_code(lot_code: str, production_date: date | None) -> bool:
    suffix = _lot_suffix(_normalized_lot_code(lot_code))
    if suffix is None:
        return False
    if production_date is not None and production_date <= CUTOFF_DATE:
        return True
    return suffix < 50


def _sample_kind(patient_id: str) -> str | None:
    patient = patient_id.strip().lower()
    if patient in CONTROL_PATIENT_IDS:
        return "control"
    for prefix in SPECIES_SHEETS:
        if patient.startswith(prefix):
            return prefix
    return None


def _load_reference_data() -> tuple[dict[str, dict], dict[str, dict[str, float | None]]]:
    import sqlite3

    conn = sqlite3.connect(SPEC_DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        specs = {}
        for row in conn.execute(
            """SELECT marker, tea, single_cv, spec_l1_od, spec_l2_od,
                      spec_n1_od, merge_bias, merge_cv
               FROM bead_ipqc_spec WHERE source = 'Qbi'"""
        ):
            specs[row["marker"].upper()] = dict(row)

        assignments = {}
        for row in conn.execute(
            """SELECT Marker, L1_89751, L2_89752, N1_45981, N3_45983
               FROM csassign"""
        ):
            assignments[row["Marker"].upper()] = {
                column: _number(row[column])
                for column in ("L1_89751", "L2_89752", "N1_45981", "N3_45983")
            }
        return specs, assignments
    finally:
        conn.close()


def _fetch_rds_records(group_key: str = "", panel_name: str = "") -> list[dict[str, Any]]:
    pg = _get_conn()
    try:
        cur = pg.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        filters = []
        params: list[Any] = []
        if group_key:
            filters.append(
                """AND (
                    REPLACE(TRIM(lot_code), '_', '') = %s
                    OR (LENGTH(REPLACE(TRIM(lot_code), '_', '')) = 12
                        AND SUBSTRING(REPLACE(TRIM(lot_code), '_', '') FROM 2) = %s)
                )"""
            )
            params.extend([group_key, group_key])
        if panel_name:
            filters.append("AND COALESCE(panel_name, '') = %s")
            params.append(panel_name)
        cur.execute(
            f"""
            SELECT id, work_order_no, lot_no, device_sn, panel_name,
                   analyze_date, analyze_time, sample_type, species, patient_id,
                   lot_code, mfg_lot_no, analyze_item, analyze_result, unit,
                   test_zone, test_well, final_delta_od, cal_od, baseline
            FROM panel_production.assay_process_records
            WHERE COALESCE(TRIM(lot_code), '') <> ''
              AND COALESCE(TRIM(analyze_item), '') <> ''
              AND (
                LOWER(COALESCE(patient_id, '')) IN ('control-1','control-2','control-3','control-4')
                OR LOWER(COALESCE(patient_id, '')) LIKE 'canine%%'
                OR LOWER(COALESCE(patient_id, '')) LIKE 'feline%%'
                OR LOWER(COALESCE(patient_id, '')) LIKE 'equine%%'
              )
              {' '.join(filters)}
            ORDER BY analyze_date DESC, analyze_time DESC, lot_code, panel_name, patient_id, analyze_item, test_well
            """,
            params,
        )
        return [dict(row) for row in cur.fetchall()]
    finally:
        pg.close()


def _build_groups(records: list[dict[str, Any]]) -> dict[tuple[str, str, str], dict[str, Any]]:
    groups: dict[tuple[str, str, str], dict[str, Any]] = {}
    for record in records:
        lot_code = _normalized_lot_code(record.get("lot_code"))
        production_date = _production_date_from_lot_code(lot_code, _clean_text(record.get("analyze_date")))
        if not _include_lot_code(lot_code, production_date):
            continue
        kind = _sample_kind(_clean_text(record.get("patient_id")))
        if not kind:
            continue

        key = (_lot_group_key(lot_code), _clean_text(record.get("panel_name")), kind)
        group = groups.setdefault(
            key,
            {
                "group_key": key[0],
                "panel_name": key[1],
                "kind": kind,
                "lot_codes": set(),
                "production_dates": [],
                "analyze_dates": [],
                "records": [],
            },
        )
        group["lot_codes"].add(lot_code)
        if production_date:
            group["production_dates"].append(production_date)
        group["analyze_dates"].append(_clean_text(record.get("analyze_date")))
        group["records"].append(record)

    for group in groups.values():
        group["lot_codes"] = sorted(group["lot_codes"])
        group["display_lot_code"] = " & ".join(group["lot_codes"])
        group["production_date"] = min(group["production_dates"]).isoformat() if group["production_dates"] else ""
        group["analyze_date"] = max(group["analyze_dates"]) if group["analyze_dates"] else ""
    return groups


def _tea_delta(spec_text: str | None, assigned: float | None) -> float | None:
    if not spec_text:
        return None
    deltas = []
    for match in re.finditer(r"([\d.]+)\s*(%)?", spec_text):
        value = _number(match.group(1))
        if value is None:
            continue
        if match.group(2) == "%":
            if assigned is not None:
                deltas.append(abs(assigned) * value / 100)
        else:
            deltas.append(value)
    return max(deltas) if deltas else None


def _assignment_for_marker(marker: str, spec: dict | None, assignments: dict[str, dict[str, float | None]]) -> dict[str, float | None]:
    direct = assignments.get(marker.upper())
    if direct:
        return direct
    spec_marker = _clean_text((spec or {}).get("marker")).upper()
    if spec_marker.startswith("Q"):
        base = re.split(r"[-_]", spec_marker[1:], maxsplit=1)[0]
        mapped = assignments.get(base.upper())
        if mapped:
            return mapped
    return {}


def _percent_text(value: float | None) -> str | None:
    if value is None:
        return None
    return f"{value:g}%"


def _stats(values: list[float], assigned: float | None = None) -> dict[str, float | None]:
    mean_value = _mean(values)
    cv_value = _cv(values)
    bias_value = (
        (mean_value - assigned) / assigned * 100
        if mean_value is not None and assigned not in (None, 0)
        else None
    )
    return {
        "mean": _round(mean_value),
        "cv": _round(cv_value, 2),
        "bias": _round(bias_value, 2),
    }


def _group_marker_summaries(group: dict[str, Any], specs: dict[str, dict], assignments: dict[str, dict[str, float | None]]) -> dict[str, dict[str, Any]]:
    summaries: dict[str, dict[str, Any]] = {}
    grouped_values: dict[tuple[str, str], list[float]] = defaultdict(list)
    grouped_ods: dict[tuple[str, str], list[float]] = defaultdict(list)
    raw_values: dict[tuple[str, str], list[tuple[str, float]]] = defaultdict(list)

    for rec in group["records"]:
        marker = _clean_text(rec.get("analyze_item"))
        patient = _clean_text(rec.get("patient_id"))
        result = _number(rec.get("analyze_result"))
        od = _number(rec.get("final_delta_od"))
        if result is not None:
            grouped_values[(patient, marker)].append(result)
            label = _clean_text(rec.get("test_zone")) or _clean_text(rec.get("test_well")) or _clean_text(rec.get("device_sn"))
            raw_values[(patient, marker)].append((label, result))
        if od is not None:
            grouped_ods[(patient, marker)].append(od)

    markers = sorted({key[1] for key in grouped_values.keys()})
    for marker in markers:
        spec = _find_spec(marker, specs)
        tea_text = spec.get("tea") if spec else None
        assigned = _assignment_for_marker(marker, spec, assignments)
        marker_summary = {
            "spec": spec or {},
            "tea_limit": tea_text,
            "controls": {},
            "samples": {},
        }
        for patient_id, (level, assignment_column) in CONTROL_LEVELS.items():
            values = grouped_values.get((patient_id, marker), [])
            assigned_value = assigned.get(assignment_column)
            tea_delta = _tea_delta(tea_text, assigned_value)
            stats = _stats(values, assigned_value)
            od_values = grouped_ods.get((patient_id, marker), [])
            stats.update(
                {
                    "assigned": assigned_value,
                    "upper": assigned_value + tea_delta if assigned_value is not None and tea_delta is not None else None,
                    "lower": assigned_value - tea_delta if assigned_value is not None and tea_delta is not None else None,
                    "values": [value for _, value in raw_values.get((patient_id, marker), [])],
                    "od_mean": _round(_mean(od_values), 4),
                    "od_cv": _round(_cv(od_values), 2),
                    "n": len(values),
                }
            )
            marker_summary["controls"][patient_id] = stats

        for (patient, sample_marker), values in grouped_values.items():
            if sample_marker == marker and patient.lower() not in CONTROL_PATIENT_IDS:
                od_values = grouped_ods.get((patient, marker), [])
                marker_summary["samples"][patient] = {
                    **_stats(values),
                    "values": [value for _, value in raw_values.get((patient, marker), [])],
                    "od_mean": _round(_mean(od_values), 4),
                    "od_cv": _round(_cv(od_values), 2),
                    "n": len(values),
                }
        summaries[marker] = marker_summary
    return summaries


def _write_cell(ws, row: int, col: int, value: Any) -> None:
    if value is None:
        return
    ws.cell(row=row, column=col, value=value)


def _clear_sheet_values(ws, rows: range, cols: range) -> None:
    for row in rows:
        for col in cols:
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if cell.data_type != "f":
                cell.value = None


def _write_all_batch_sheet(ws, groups: list[dict[str, Any]], summaries_by_group: dict[str, dict[str, dict[str, Any]]]) -> int:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    row = 2
    for group in groups:
        summaries = summaries_by_group[group["id"]]
        for marker, summary in summaries.items():
            ws.cell(row, 1, group["production_date"])
            ws.cell(row, 2, group["panel_name"])
            ws.cell(row, 3, group["display_lot_code"])
            ws.cell(row, 4, marker)
            spec = summary["spec"]
            tea = summary["tea_limit"]
            for patient_id, config in RESULT_COLUMNS.items():
                control = summary["controls"].get(patient_id, {})
                assigned = control.get("assigned")
                _write_cell(ws, row, config["tea"], tea)
                _write_cell(ws, row, config["assigned"], assigned)
                _write_cell(ws, row, config["upper"], control.get("upper"))
                _write_cell(ws, row, config["lower"], control.get("lower"))
                for col, value in zip(config["values"], control.get("values", [])):
                    _write_cell(ws, row, col, value)
                _write_cell(ws, row, config["mean"], control.get("mean"))
                _write_cell(ws, row, config["bias"], control.get("bias"))
                _write_cell(ws, row, config["cv"], _percent_text(control.get("cv")))
            row += 1
    return row - 2


def _empty_marker_summary() -> dict[str, Any]:
    return {"spec": {}, "tea_limit": None, "controls": {}, "samples": {}}


def _write_summary_sheet(ws, group: dict[str, Any] | None, summaries: dict[str, dict[str, Any]], sheet_kind: str, markers: list[str] | None = None) -> None:
    _clear_sheet_values(ws, range(5, ws.max_row + 1), range(4, min(ws.max_column, 36) + 1))
    if sheet_kind != "control":
        _clear_sheet_values(ws, range(DETAIL_START_ROW, ws.max_row + 1), range(2, min(ws.max_column, 37) + 1))

    markers = (markers or list(summaries.keys()))[:10]
    if not markers:
        return
    for idx, marker in enumerate(markers):
        marker_col = SUMMARY_MARKER_COLUMNS[idx]
        od_col = SUMMARY_OD_COLUMNS[idx]
        ws.cell(3, marker_col, marker)
        ws.cell(4, marker_col, f"{marker}(原線)")
        ws.cell(4, marker_col + 1, f"{marker}\n換線後")
        ws.cell(3, od_col, marker)
        ws.cell(34, marker_col, marker)
        ws.cell(35, marker_col, f"{marker}(原線)")
        ws.cell(35, marker_col + 1, f"{marker}\n換線後")
        ws.cell(34, od_col, marker)

    if sheet_kind == "control":
        for idx, marker in enumerate(markers):
            summary = summaries.get(marker, _empty_marker_summary())
            marker_col = SUMMARY_MARKER_COLUMNS[idx]
            od_col = SUMMARY_OD_COLUMNS[idx]
            for patient_id, rows in CONTROL_SUMMARY_ROWS.items():
                control = summary["controls"].get(patient_id, {})
                ws.cell(5, marker_col, summary.get("tea_limit"))
                ws.cell(6, marker_col, control.get("assigned"))
                ws.cell(7, marker_col, control.get("lower"))
                ws.cell(8, marker_col, control.get("upper"))
                mean_row, bias_row, cv_row = rows
                ws.cell(mean_row, marker_col, control.get("mean"))
                ws.cell(bias_row, marker_col, control.get("bias"))
                ws.cell(cv_row, marker_col, _percent_text(control.get("cv")))
                ws.cell(mean_row, od_col, control.get("od_mean"))
    else:
        prefix = sheet_kind
        patients = sorted(
            {
                patient
                for marker_summary in summaries.values()
                for patient in marker_summary["samples"]
                if patient.lower().startswith(prefix)
            }
        )[:7]
        for idx, marker in enumerate(markers):
            marker_col = SUMMARY_MARKER_COLUMNS[idx]
            od_col = SUMMARY_OD_COLUMNS[idx]
            for p_index, patient in enumerate(patients):
                rows = (9 + p_index * 3, 10 + p_index * 3, 11 + p_index * 3)
                sample = summaries.get(marker, _empty_marker_summary())["samples"].get(patient, {})
                ws.cell(rows[0], marker_col, sample.get("mean"))
                ws.cell(rows[1], marker_col, sample.get("bias"))
                ws.cell(rows[2], marker_col, _percent_text(sample.get("cv")))
                ws.cell(rows[0], od_col, sample.get("od_mean"))

    if sheet_kind != "control" or not group:
        return

    detail_rows = []
    for rec in group["records"]:
        if sheet_kind == "control" and _clean_text(rec.get("patient_id")).lower() not in CONTROL_PATIENT_IDS:
            continue
        if sheet_kind != "control" and not _clean_text(rec.get("patient_id")).lower().startswith(sheet_kind):
            continue
        detail_rows.append(rec)

    for offset, rec in enumerate(detail_rows[:DETAIL_MAX_ROWS]):
        row = DETAIL_START_ROW + offset
        ws.cell(row, 2, _clean_text(rec.get("patient_id")))
        ws.cell(row, 3, _clean_text(rec.get("device_sn")))
        ws.cell(row, 4, _clean_text(rec.get("test_zone")) or _clean_text(rec.get("test_well")))
        marker = _clean_text(rec.get("analyze_item"))
        if marker in markers:
            idx = markers.index(marker)
            result = _number(rec.get("analyze_result"))
            od = _number(rec.get("final_delta_od"))
            ws.cell(row, SUMMARY_MARKER_COLUMNS[idx], result)
            ws.cell(row, SUMMARY_OD_COLUMNS[idx], od)
        ws.cell(row, 37, _clean_text(rec.get("device_sn")))


def _first_non_empty(records: list[dict[str, Any]], field: str) -> str:
    for rec in records:
        value = _clean_text(rec.get(field))
        if value:
            return value
    return ""


def _fetch_product_code(panel_name: str, group_key: str) -> str:
    sub_panel = group_key[:3] if group_key else ""
    pg = _get_conn()
    try:
        cur = pg.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            """
            SELECT product_code
            FROM qbi_qr.panels
            WHERE panel_name = %s
            ORDER BY CASE
                WHEN panel_key = %s THEN 0
                WHEN sub_panel_type = %s THEN 1
                WHEN panel_key LIKE %s THEN 2
                ELSE 3
            END
            LIMIT 1
            """,
            (panel_name, f"00-{sub_panel}", sub_panel, f"%{sub_panel}"),
        )
        row = cur.fetchone()
        return _clean_text(row.get("product_code") if row else "")
    finally:
        pg.close()


def _fetch_work_order_info(group: dict[str, Any] | None) -> dict[str, Any]:
    if not group:
        return {"work_order_no": "", "lot_no": "", "form_data": None, "source_table": ""}
    records = group.get("records", [])
    lot_candidates = []
    work_order_candidates = []
    for rec in records:
        for field in ("mfg_lot_no", "lot_no"):
            value = _clean_text(rec.get(field))
            if value and value not in lot_candidates:
                lot_candidates.append(value)
        value = _clean_text(rec.get("work_order_no"))
        if value and value not in work_order_candidates:
            work_order_candidates.append(value)
    if not lot_candidates and not work_order_candidates:
        return {"work_order_no": "", "lot_no": "", "form_data": None, "source_table": ""}

    pg = _get_conn()
    try:
        cur = pg.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        params = [lot_candidates, work_order_candidates, lot_candidates, work_order_candidates]
        cur.execute(
            """
            SELECT source_table, work_order_no, lot_no, form_data, created_at, updated_at
            FROM (
                SELECT 'tutti_work_orders' AS source_table, work_order_no, lot_no, form_data, created_at, updated_at
                FROM panel_production.tutti_work_orders
                WHERE lot_no = ANY(%s) OR work_order_no = ANY(%s)
                UNION ALL
                SELECT 'tutti_work_orders_water' AS source_table, work_order_no, lot_no, form_data, created_at, updated_at
                FROM panel_production.tutti_work_orders_water
                WHERE lot_no = ANY(%s) OR work_order_no = ANY(%s)
            ) q
            ORDER BY updated_at DESC NULLS LAST, created_at DESC NULLS LAST
            LIMIT 1
            """,
            params,
        )
        row = cur.fetchone()
        if not row:
            return {"work_order_no": "", "lot_no": lot_candidates[0] if lot_candidates else "", "form_data": None, "source_table": ""}
        form_data = row.get("form_data")
        if isinstance(form_data, str):
            try:
                form_data = json.loads(form_data)
            except json.JSONDecodeError:
                form_data = None
        return {
            "work_order_no": _clean_text(row.get("work_order_no")),
            "lot_no": _clean_text(row.get("lot_no")),
            "form_data": form_data if isinstance(form_data, dict) else None,
            "source_table": _clean_text(row.get("source_table")),
        }
    finally:
        pg.close()


def _marker_batch_text(batch_info: dict[str, str]) -> str:
    parts = []
    for label, key in (("d", "d_lot"), ("D", "bigD_lot"), ("U", "u_lot")):
        value = _clean_text(batch_info.get(key))
        if value:
            parts.append(f"{label}:{value}")
    return " / ".join(parts)


def _cs_real_context(dataset: dict[str, Any], group: dict[str, Any] | None, markers: list[str]) -> dict[str, Any]:
    records = group.get("records", []) if group else []
    work_order = _fetch_work_order_info(group)
    mfg_lot_no = _first_non_empty(records, "mfg_lot_no") or work_order.get("lot_no", "") or _first_non_empty(records, "lot_code")
    product_code = _fetch_product_code(dataset.get("panel_name", ""), dataset.get("group_key", ""))
    test_times = sorted({_clean_text(rec.get("analyze_time")) for rec in records if _clean_text(rec.get("analyze_time"))})
    device_sns = sorted({_clean_text(rec.get("device_sn")) for rec in records if _clean_text(rec.get("device_sn"))})
    batch_lookup = _fetch_batch_info(mfg_lot_no, dataset.get("panel_name", ""), dataset.get("analyze_date", ""), markers)
    maker_batches = [_marker_batch_text(batch_lookup.get(marker, batch_lookup.get(marker.upper(), {}))) for marker in markers]
    return {
        "page_info": [
            {"label": "Panel Name", "value": dataset.get("panel_name", "")},
            {"label": "Product Code (REF)", "value": product_code},
            {"label": "Lot No.", "value": dataset.get("display_lot_code", "")},
            {"label": "Work Order", "value": work_order.get("work_order_no", "")},
            {"label": "MFG Lot No.", "value": work_order.get("lot_no") or mfg_lot_no},
            {"label": "Production Date", "value": dataset.get("production_date", "")},
            {"label": "Test Date", "value": dataset.get("analyze_date", "")},
            {"label": "Test Time", "value": " / ".join(test_times[:3])},
            {"label": "Device SN", "value": " / ".join(device_sns[:3])},
            {"label": "Maker Batch Source", "value": "baseline_service._fetch_batch_info"},
        ],
        "maker_batch": {"markers": markers, "values": maker_batches},
    }


def _write_cs_real(ws, dataset: dict[str, Any], group: dict[str, Any] | None, summaries: dict[str, dict[str, Any]], markers: list[str] | None = None) -> None:
    _clear_sheet_values(ws, range(2, ws.max_row + 1), range(4, min(ws.max_column, 35) + 1))
    if not group:
        return
    markers = (markers or list(summaries.keys()))[:10]
    context = _cs_real_context(dataset, group, markers)
    info = {item["label"]: item["value"] for item in context["page_info"]}
    ws.cell(2, 4, info.get("Panel Name"))
    ws.cell(2, 14, info.get("Work Order"))
    ws.cell(3, 4, info.get("Product Code (REF)"))
    ws.cell(3, 14, info.get("Production Date"))
    ws.cell(4, 4, info.get("Lot No."))
    ws.cell(4, 8, info.get("MFG Lot No."))
    ws.cell(4, 14, info.get("Test Date"))
    ws.cell(5, 4, info.get("Device SN"))
    ws.cell(5, 14, info.get("Test Time"))
    for idx, marker in enumerate(markers):
        col = 4 + idx * 2
        od_col = 26 + idx
        ws.cell(20, col, marker)
        ws.cell(20, col + 1, f"{marker}\n換線後")
        ws.cell(20, od_col, marker)
        ws.cell(21, col, context["maker_batch"]["values"][idx] if idx < len(context["maker_batch"]["values"]) else None)
        summary = summaries.get(marker, _empty_marker_summary())
        for patient_id, rows in CONTROL_SUMMARY_ROWS.items():
            control = summary["controls"].get(patient_id, {})
            row_offset = {"Control-1": 24, "Control-2": 27, "Control-3": 30, "Control-4": 33}[patient_id]
            ws.cell(row_offset, col, control.get("mean"))
            ws.cell(row_offset + 1, col, control.get("bias"))
            ws.cell(row_offset + 2, col, _percent_text(control.get("cv")))
            ws.cell(row_offset, od_col, control.get("od_mean"))

def _dataset_id(group_key: str, panel_name: str) -> str:
    return f"{group_key}::{panel_name}"


def _collect_datasets(groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    datasets: dict[str, dict[str, Any]] = {}
    for group in groups:
        dataset_id = _dataset_id(group["group_key"], group["panel_name"])
        dataset = datasets.setdefault(
            dataset_id,
            {
                "id": dataset_id,
                "group_key": group["group_key"],
                "panel_name": group["panel_name"],
                "lot_codes": set(),
                "production_dates": [],
                "analyze_dates": [],
                "groups_by_kind": {},
            },
        )
        dataset["groups_by_kind"][group["kind"]] = group
        dataset["lot_codes"].update(group["lot_codes"])
        if group.get("production_date"):
            dataset["production_dates"].append(group["production_date"])
        if group.get("analyze_date"):
            dataset["analyze_dates"].append(group["analyze_date"])

    result = []
    for dataset in datasets.values():
        dataset["lot_codes"] = sorted(dataset["lot_codes"])
        dataset["display_lot_code"] = " & ".join(dataset["lot_codes"])
        dataset["production_date"] = min(dataset["production_dates"]) if dataset["production_dates"] else ""
        dataset["analyze_date"] = max(dataset["analyze_dates"]) if dataset["analyze_dates"] else ""
        dataset["kinds"] = sorted(dataset["groups_by_kind"].keys())
        dataset["record_count"] = sum(len(group["records"]) for group in dataset["groups_by_kind"].values())
        result.append(dataset)
    result.sort(key=lambda row: (row["production_date"], row["display_lot_code"], row["panel_name"]), reverse=True)
    return result


def _summaries_for_dataset(dataset: dict[str, Any], specs: dict[str, dict], assignments: dict[str, dict[str, float | None]]) -> dict[str, dict[str, dict[str, Any]]]:
    return {
        kind: _group_marker_summaries(group, specs, assignments)
        for kind, group in dataset["groups_by_kind"].items()
    }


def _control_label(patient_id: str) -> str:
    patient = patient_id.strip().lower()
    if patient.startswith("control-"):
        return f"Control {patient.split('-', 1)[1]}"
    for prefix in ("canine", "feline", "equine"):
        if patient.startswith(prefix):
            suffix = patient[len(prefix):].strip(" -_")
            return f"{prefix.title()} {suffix}".strip()
    return patient_id


def _sheet_patients(group: dict[str, Any] | None, sheet_kind: str) -> list[str]:
    if not group:
        return []
    patients = []
    for rec in group["records"]:
        patient = _clean_text(rec.get("patient_id"))
        lower = patient.lower()
        if sheet_kind == "control" and lower in CONTROL_PATIENT_IDS:
            patients.append(patient)
        elif sheet_kind != "control" and lower.startswith(sheet_kind):
            patients.append(patient)
    return sorted(set(patients), key=lambda p: (p.lower().split("-")[-1], p.lower()))


def _summary_table(markers: list[str], summaries: dict[str, dict[str, Any]], sheet_kind: str, metric: str, group: dict[str, Any] | None) -> dict[str, Any]:
    rows = []
    marker_summary = lambda marker: summaries.get(marker, _empty_marker_summary())
    if sheet_kind == "control":
        rows.extend([
            {"label": "", "stat": "TEa", "values": [marker_summary(marker).get("tea_limit") for marker in markers]},
            {"label": "", "stat": "Assign value" if metric == "conc" else "Assay Value", "values": [
                marker_summary(marker)["controls"].get("Control-1", {}).get("assigned") if metric == "conc" else None
                for marker in markers
            ]},
            {"label": "", "stat": "LCL", "values": [
                marker_summary(marker)["controls"].get("Control-1", {}).get("lower") if metric == "conc" else None
                for marker in markers
            ]},
            {"label": "", "stat": "UCL", "values": [
                marker_summary(marker)["controls"].get("Control-1", {}).get("upper") if metric == "conc" else None
                for marker in markers
            ]},
        ])
        stat_fields = (("Mean", "mean"), ("Bias", "bias"), ("CV%", "cv")) if metric == "conc" else (("Mean", "od_mean"), ("CV%", "od_cv"))
        for patient_id in ("Control-1", "Control-2", "Control-3", "Control-4"):
            for stat, field in stat_fields:
                values = []
                for marker in markers:
                    control = marker_summary(marker)["controls"].get(patient_id, {})
                    values.append(control.get(field))
                rows.append({"label": patient_id.replace("-", " "), "stat": stat, "values": values})
    else:
        rows.extend([
            {"label": "", "stat": "TEa", "values": [marker_summary(marker).get("tea_limit") for marker in markers]},
            {"label": "", "stat": "Assign value" if metric == "conc" else "Assay Value", "values": [None for _ in markers]},
            {"label": "", "stat": "LCL", "values": [None for _ in markers]},
            {"label": "", "stat": "UCL", "values": [None for _ in markers]},
        ])
        stat_fields = (("Mean", "mean"), ("Bias", "bias"), ("CV%", "cv")) if metric == "conc" else (("Mean", "od_mean"), ("CV%", "od_cv"))
        for patient in _sheet_patients(group, sheet_kind):
            for stat, field in stat_fields:
                values = []
                for marker in markers:
                    sample = summaries.get(marker, _empty_marker_summary())["samples"].get(patient, {})
                    values.append(sample.get(field))
                rows.append({"label": _control_label(patient), "stat": stat, "values": values})
    return {"markers": markers, "rows": rows}


def _detail_key(rec: dict[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        _clean_text(rec.get("patient_id")),
        _clean_text(rec.get("device_sn")),
        _clean_text(rec.get("test_zone")) or _clean_text(rec.get("test_well")),
        _clean_text(rec.get("analyze_date")),
        _clean_text(rec.get("analyze_time")),
    )


def _test_record_count(group: dict[str, Any] | None, sheet_kind: str) -> int:
    keys = set()
    if not group:
        return 0
    for rec in group["records"]:
        patient = _clean_text(rec.get("patient_id"))
        lower = patient.lower()
        if sheet_kind == "control" and lower not in CONTROL_PATIENT_IDS:
            continue
        if sheet_kind != "control" and not lower.startswith(sheet_kind):
            continue
        keys.add(_detail_key(rec))
    return len(keys)

def _detail_table(group: dict[str, Any] | None, markers: list[str], sheet_kind: str, metric: str) -> dict[str, Any]:
    grouped: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    if group:
        for rec in group["records"]:
            patient = _clean_text(rec.get("patient_id"))
            lower = patient.lower()
            if sheet_kind == "control" and lower not in CONTROL_PATIENT_IDS:
                continue
            if sheet_kind != "control" and not lower.startswith(sheet_kind):
                continue
            key = _detail_key(rec)
            row = grouped.setdefault(key, {
                "sample": _control_label(patient),
                "device_sn": _clean_text(rec.get("device_sn")),
                "test_zone": _clean_text(rec.get("test_zone")) or _clean_text(rec.get("test_well")),
                "analyze_time": _clean_text(rec.get("analyze_time")),
                "values": {marker: {"original": None, "changed": None} for marker in markers},
            })
            marker = _clean_text(rec.get("analyze_item"))
            if marker not in markers:
                continue
            value = _number(rec.get("analyze_result")) if metric == "conc" else _number(rec.get("final_delta_od"))
            row["values"][marker]["original"] = value
    rows = list(grouped.values())
    rows.sort(key=lambda row: (row["sample"], row["device_sn"], row["test_zone"], row["analyze_time"]))
    return {"markers": markers, "rows": rows[:DETAIL_MAX_ROWS], "value_mode": metric}


def _sheet_preview(group: dict[str, Any] | None, summaries: dict[str, dict[str, Any]], sheet_kind: str, markers: list[str] | None = None) -> dict[str, Any]:
    sheet_markers = (markers or list(summaries.keys()))[:10]
    preview = {
        "markers": sheet_markers,
        "test_count": _test_record_count(group, sheet_kind),
        "summary_conc": _summary_table(sheet_markers, summaries, sheet_kind, "conc", group),
        "summary_od": _summary_table(sheet_markers, summaries, sheet_kind, "od", group),
    }
    if sheet_kind == "control":
        preview["detail_conc"] = _detail_table(group, sheet_markers, sheet_kind, "conc")
        preview["detail_od"] = _detail_table(group, sheet_markers, sheet_kind, "od")
    return preview


def _sheet_preview_rows(group: dict[str, Any] | None, summaries: dict[str, dict[str, Any]], sheet_kind: str) -> list[dict[str, Any]]:
    legacy = []
    if not group:
        return legacy
    markers = list(summaries.keys())[:10]
    for marker in markers:
        summary = summaries[marker]
        if sheet_kind == "control":
            for patient_id in ("Control-1", "Control-2", "Control-3", "Control-4"):
                control = summary["controls"].get(patient_id, {})
                legacy.append({
                    "type": patient_id,
                    "marker": marker,
                    "tea": summary.get("tea_limit"),
                    "assigned": control.get("assigned"),
                    "lcl": control.get("lower"),
                    "ucl": control.get("upper"),
                    "mean": control.get("mean"),
                    "bias": control.get("bias"),
                    "cv": control.get("cv"),
                    "od_mean": control.get("od_mean"),
                    "n": control.get("n"),
                })
        else:
            for patient, sample in sorted(summary["samples"].items()):
                if not patient.lower().startswith(sheet_kind):
                    continue
                legacy.append({
                    "type": patient,
                    "marker": marker,
                    "mean": sample.get("mean"),
                    "bias": sample.get("bias"),
                    "cv": sample.get("cv"),
                    "od_mean": sample.get("od_mean"),
                    "n": sample.get("n"),
                })
    return legacy

def _report_markers(summaries_by_kind: dict[str, dict[str, dict[str, Any]]]) -> list[str]:
    control_markers = list(summaries_by_kind.get("control", {}).keys())
    if control_markers:
        return control_markers[:10]
    markers = sorted({marker for summary in summaries_by_kind.values() for marker in summary.keys()})
    return markers[:10]

def _cs_real_preview(dataset: dict[str, Any], group: dict[str, Any] | None, summaries: dict[str, dict[str, Any]], markers: list[str]) -> dict[str, Any]:
    context = _cs_real_context(dataset, group, markers)
    conc = _summary_table(markers, summaries, "control", "conc", group)
    od = _summary_table(markers, summaries, "control", "od", group)
    return {
        "sheet_name": "CS-real彙總",
        "markers": markers,
        "test_count": _test_record_count(group, "control"),
        "page_info": context["page_info"],
        "maker_batch": context["maker_batch"],
        "summary_conc": conc,
        "summary_od": od,
        "rows": [],
    }


def _preview_for_dataset(report_path: Path, dataset: dict[str, Any], summaries_by_kind: dict[str, dict[str, dict[str, Any]]]) -> dict[str, Any]:
    sheets = {}
    report_markers = _report_markers(summaries_by_kind)
    control_group = dataset["groups_by_kind"].get("control")
    control_summary = summaries_by_kind.get("control", {})
    sheets["CS-real彙總"] = _cs_real_preview(dataset, control_group, control_summary, report_markers)
    for kind, sheet_name in {"control": "Control", **SPECIES_SHEETS}.items():
        group = dataset["groups_by_kind"].get(kind)
        summary = summaries_by_kind.get(kind, {})
        sheet_preview = _sheet_preview(group, summary, kind, report_markers)
        sheet_preview["sheet_name"] = sheet_name
        sheet_preview["rows"] = _sheet_preview_rows(group, summary, kind)
        sheets[sheet_name] = sheet_preview
    markers = report_markers
    return {
        "ok": True,
        "file_name": report_path.name,
        "download_url": f"/api/assayprocess/lot-reports/{report_path.name}/download",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "id": dataset["id"],
        "panel_name": dataset["panel_name"],
        "display_lot_code": dataset["display_lot_code"],
        "lot_codes": dataset["lot_codes"],
        "production_date": dataset["production_date"],
        "analyze_date": dataset["analyze_date"],
        "record_count": dataset["record_count"],
        "kinds": dataset["kinds"],
        "sheets": sheets,
        "groups": [{
            "id": dataset["id"],
            "panel_name": dataset["panel_name"],
            "kind": ",".join(dataset["kinds"]),
            "display_lot_code": dataset["display_lot_code"],
            "lot_codes": dataset["lot_codes"],
            "production_date": dataset["production_date"],
            "analyze_date": dataset["analyze_date"],
            "record_count": dataset["record_count"],
            "marker_count": len(markers),
            "markers": markers,
        }],
    }



def _json_preview(report_path: Path, groups: list[dict[str, Any]], summaries_by_group: dict[str, dict[str, dict[str, Any]]], all_batch_rows: int) -> dict[str, Any]:
    preview_groups = []
    for group in groups[:100]:
        summaries = summaries_by_group[group["id"]]
        preview_groups.append(
            {
                "id": group["id"],
                "panel_name": group["panel_name"],
                "kind": group["kind"],
                "display_lot_code": group["display_lot_code"],
                "lot_codes": group["lot_codes"],
                "production_date": group["production_date"],
                "analyze_date": group["analyze_date"],
                "record_count": len(group["records"]),
                "marker_count": len(summaries),
                "markers": list(summaries.keys())[:20],
            }
        )
    return {
        "ok": True,
        "file_name": report_path.name,
        "download_url": f"/api/assayprocess/lot-reports/{report_path.name}/download",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "all_batch_rows": all_batch_rows,
        "group_count": len(groups),
        "groups": preview_groups,
    }


def _load_datasets() -> tuple[list[dict[str, Any]], dict[str, dict], dict[str, dict[str, float | None]]]:
    specs, assignments = _load_reference_data()
    records = _fetch_rds_records()
    groups_map = _build_groups(records)
    groups = []
    for index, group in enumerate(groups_map.values()):
        group["id"] = f"{group['group_key']}::{group['panel_name']}::{group['kind']}::{index}"
        groups.append(group)
    return _collect_datasets(groups), specs, assignments


def _dataset_id_parts(dataset_id: str) -> tuple[str, str]:
    if "::" not in dataset_id:
        return "", ""
    group_key, panel_name = dataset_id.split("::", 1)
    return group_key.strip(), panel_name.strip()


def _load_selected_dataset(dataset_id: str = "", lot_code: str = "") -> tuple[dict[str, Any] | None, dict[str, dict], dict[str, dict[str, float | None]]]:
    specs, assignments = _load_reference_data()
    group_key, panel_name = _dataset_id_parts(dataset_id)
    if not group_key and lot_code:
        group_key = _lot_group_key(lot_code)
    if not group_key:
        return None, specs, assignments
    records = _fetch_rds_records(group_key=group_key, panel_name=panel_name)
    groups_map = _build_groups(records)
    groups = []
    for index, group in enumerate(groups_map.values()):
        group["id"] = f"{group['group_key']}::{group['panel_name']}::{group['kind']}::{index}"
        groups.append(group)
    datasets = _collect_datasets(groups)
    return _select_dataset(datasets, dataset_id=dataset_id, lot_code=lot_code), specs, assignments

def list_lot_report_groups() -> dict[str, Any]:
    datasets, _specs, _assignments = _load_datasets()
    rows = [
        {
            "id": dataset["id"],
            "panel_name": dataset["panel_name"],
            "display_lot_code": dataset["display_lot_code"],
            "lot_codes": dataset["lot_codes"],
            "production_date": dataset["production_date"],
            "analyze_date": dataset["analyze_date"],
            "record_count": dataset["record_count"],
            "kinds": dataset["kinds"],
        }
        for dataset in datasets
    ]
    return {"ok": True, "rows": rows, "total": len(rows)}


def list_lot_reports() -> dict[str, Any]:
    reports = []
    for path in sorted(EXCEL_DATA_DIR.glob("*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True):
        if path.parent.name == "templet":
            continue
        sidecar = path.with_suffix(".json")
        reports.append(
            {
                "file_name": path.name,
                "size": path.stat().st_size,
                "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                "download_url": f"/api/assayprocess/lot-reports/{path.name}/download",
                "has_preview": sidecar.exists(),
            }
        )
    return {"ok": True, "reports": reports}


def get_lot_report_preview(file_name: str) -> dict[str, Any]:
    safe_name = Path(file_name).name
    sidecar = (EXCEL_DATA_DIR / safe_name).with_suffix(".json")
    if not sidecar.exists():
        return {"ok": False, "error": "preview not found"}
    return json.loads(sidecar.read_text(encoding="utf-8"))


def _safe_report_stem(display_lot_code: str) -> str:
    stem = re.sub(r"\s*&\s*", "&", display_lot_code.strip())
    stem = re.sub(r"[^0-9A-Za-z&_-]+", "_", stem)
    return stem.strip("_") or "lot_code"


def _select_dataset(datasets: list[dict[str, Any]], dataset_id: str = "", lot_code: str = "") -> dict[str, Any] | None:
    normalized_lot = _normalized_lot_code(lot_code)
    for dataset in datasets:
        if dataset_id and dataset["id"] == dataset_id:
            return dataset
        if normalized_lot and normalized_lot in dataset["lot_codes"]:
            return dataset
    return None


def generate_lot_report(output_date: str | None = None, dataset_id: str = "", lot_code: str = "") -> dict[str, Any]:
    dataset, specs, assignments = _load_selected_dataset(dataset_id=dataset_id, lot_code=lot_code)
    if not dataset:
        datasets, specs, assignments = _load_datasets()
        dataset = _select_dataset(datasets, dataset_id=dataset_id, lot_code=lot_code)
    if not dataset:
        return {"ok": False, "error": "selected lot_code not found"}

    summaries_by_kind = _summaries_for_dataset(dataset, specs, assignments)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    control_group = dataset["groups_by_kind"].get("control")
    report_markers = _report_markers(summaries_by_kind)
    _write_summary_sheet(wb["Control"], control_group, summaries_by_kind.get("control", {}), "control", report_markers)
    for kind, sheet in SPECIES_SHEETS.items():
        group = dataset["groups_by_kind"].get(kind)
        _write_summary_sheet(wb[sheet], group, summaries_by_kind.get(kind, {}), kind, report_markers)

    _write_cs_real(wb["CS-real彙整"], dataset, control_group, summaries_by_kind.get("control", {}), report_markers)

    stamp = output_date or datetime.now().strftime("%y%m%d")
    stem = _safe_report_stem(dataset["display_lot_code"])
    report_path = EXCEL_DATA_DIR / f"{stem}_{stamp}.xlsx"
    wb.save(report_path)

    preview = _preview_for_dataset(report_path, dataset, summaries_by_kind)
    report_path.with_suffix(".json").write_text(json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8")
    return preview


def report_file_path(file_name: str) -> Path:
    safe_name = Path(file_name).name
    path = EXCEL_DATA_DIR / safe_name
    if path.suffix.lower() != ".xlsx" or safe_name == TEMPLATE_PATH.name:
        raise FileNotFoundError("invalid report file")
    if not path.exists():
        raise FileNotFoundError(safe_name)
    return path
